"""Manage the Excel tracker (data/tracker.xlsx)."""

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from config import TRACKER_PATH

logger = logging.getLogger(__name__)

HEADERS = [
    "ID",
    "Date Found",
    "Source",
    "Company",
    "Job Title",
    "Location",
    "Region",
    "Date Posted",
    "Status",
    "Job URL",
    "Resume",
    "Cover Letter",
    "Notes",
]

STATUS_COLORS = {
    "Waiting to apply": "FFF9C4",   # pale yellow
    "Applied":          "C8E6C9",   # pale green
    "Rejected":         "FFCDD2",   # pale red
    "Interview":        "BBDEFB",   # pale blue
    "Offer":            "B2EBF2",   # pale teal
}

HEADER_FILL  = PatternFill("solid", fgColor="1565C0")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BORDER_SIDE  = Side(style="thin", color="CCCCCC")
CELL_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)
COL_WIDTHS = {
    "ID":           8,
    "Date Found":   18,
    "Source":       12,
    "Company":      22,
    "Job Title":    30,
    "Location":     22,
    "Region":       14,
    "Date Posted":  18,
    "Status":       18,
    "Job URL":      40,
    "Resume":       35,
    "Cover Letter": 35,
    "Notes":        30,
}


def _get_or_create_wb() -> tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(str(TRACKER_PATH))
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Internship Tracker"
        _write_headers(ws)
        wb.save(str(TRACKER_PATH))
    return wb, ws


def _write_headers(ws):
    ws.append(HEADERS)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 15)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _existing_ids(ws) -> set[str]:
    ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            ids.add(str(row[0]))
    return ids


def _color_status_cell(cell, status: str):
    color = STATUS_COLORS.get(status, "FFFFFF")
    cell.fill = PatternFill("solid", fgColor=color)


def add_jobs(jobs_with_docs: list[dict]) -> int:
    """
    Append new jobs to the tracker.
    `jobs_with_docs` items must have keys from `job` dict + 'resume_path' + 'cover_path'.
    Returns number of rows added.
    """
    wb, ws = _get_or_create_wb()
    existing = _existing_ids(ws)
    added = 0

    for entry in jobs_with_docs:
        job_id = str(entry.get("job_id", ""))
        if job_id in existing:
            continue

        posted_str = ""
        if entry.get("posted_at"):
            try:
                posted_str = entry["posted_at"].strftime("%Y-%m-%d %H:%M UTC")
            except Exception:
                posted_str = str(entry.get("posted_raw", ""))

        row = [
            job_id,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            entry.get("source", ""),
            entry.get("company", ""),
            entry.get("title", ""),
            entry.get("location", ""),
            entry.get("region", ""),
            posted_str,
            "Waiting to apply",
            entry.get("url", ""),
            str(entry.get("resume_path", "")),
            str(entry.get("cover_path", "")),
            "",
        ]
        ws.append(row)
        row_idx = ws.max_row

        # Style the new row
        for col_idx, _ in enumerate(HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border    = CELL_BORDER
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            cell.font      = Font(size=9)

        # Colour status cell
        status_col = HEADERS.index("Status") + 1
        _color_status_cell(ws.cell(row=row_idx, column=status_col), "Waiting to apply")

        # Make URLs clickable
        url_col = HEADERS.index("Job URL") + 1
        url_cell = ws.cell(row=row_idx, column=url_col)
        if entry.get("url"):
            url_cell.hyperlink = entry["url"]
            url_cell.font = Font(size=9, color="1565C0", underline="single")

        existing.add(job_id)
        added += 1

    if added:
        wb.save(str(TRACKER_PATH))
        logger.info("Tracker updated: +%d rows → %s", added, TRACKER_PATH)
    return added


def mark_applied(job_id: str, notes: str = "") -> bool:
    """
    Update job Status to 'Applied' in the tracker and optionally add notes.
    Returns True if the job was found and updated, False otherwise.
    """
    if not TRACKER_PATH.exists():
        return False
    try:
        wb = openpyxl.load_workbook(str(TRACKER_PATH))
        ws = wb.active
        id_col     = HEADERS.index("ID") + 1
        status_col = HEADERS.index("Status") + 1
        notes_col  = HEADERS.index("Notes") + 1

        for row in ws.iter_rows(min_row=2):
            cell_id = row[id_col - 1].value
            if cell_id is not None and str(cell_id).strip() == str(job_id).strip():
                row[status_col - 1].value = "Applied"
                _color_status_cell(row[status_col - 1], "Applied")
                if notes:
                    existing = row[notes_col - 1].value or ""
                    row[notes_col - 1].value = (
                        f"{existing}  [{notes}]".strip() if existing else notes
                    )
                wb.save(str(TRACKER_PATH))
                logger.info("Marked Applied: job_id=%s", job_id)
                return True

        logger.warning("mark_applied: job_id=%s not found in tracker", job_id)
        return False
    except Exception as exc:
        logger.error("mark_applied failed: %s", exc)
        return False


def get_all_jobs() -> list[dict]:
    """Return every tracker row as a list of dicts (keys = HEADERS)."""
    if not TRACKER_PATH.exists():
        return []
    try:
        wb = openpyxl.load_workbook(str(TRACKER_PATH), read_only=True)
        ws = wb.active
        jobs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            jobs.append(dict(zip(HEADERS, row)))
        return jobs
    except Exception as exc:
        logger.error("get_all_jobs failed: %s", exc)
        return []


def get_tracker_path() -> Path:
    return TRACKER_PATH


def update_status(job_id: str, new_status: str, notes: str = "") -> bool:
    """
    Set any status value on a tracker row.
    Accepts: Applied, Waiting to apply, Rejected, Interview, Offer.
    Returns True if found and updated.
    """
    if not TRACKER_PATH.exists():
        return False
    try:
        wb = openpyxl.load_workbook(str(TRACKER_PATH))
        ws = wb.active
        id_col     = HEADERS.index("ID") + 1
        status_col = HEADERS.index("Status") + 1
        notes_col  = HEADERS.index("Notes") + 1

        for row in ws.iter_rows(min_row=2):
            cell_id = row[id_col - 1].value
            if cell_id is not None and str(cell_id).strip() == str(job_id).strip():
                row[status_col - 1].value = new_status
                _color_status_cell(row[status_col - 1], new_status)
                if notes:
                    existing = row[notes_col - 1].value or ""
                    row[notes_col - 1].value = (
                        f"{existing}  [{notes}]".strip() if existing else notes
                    )
                wb.save(str(TRACKER_PATH))
                logger.info("Status updated: job_id=%s → %s", job_id, new_status)
                return True

        return False
    except Exception as exc:
        logger.error("update_status failed: %s", exc)
        return False


def create_stub(job_id: str, status: str = "Applied", notes: str = "") -> bool:
    """
    Create a minimal tracker row for a job_id that isn't in the tracker yet.
    Used by the deferred apply queue when a button tap fires before the main
    run has committed the full job data.
    Returns True if a new row was created (False if already exists).
    """
    wb, ws = _get_or_create_wb()
    existing = _existing_ids(ws)
    if job_id in existing:
        return False

    row = [
        job_id,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        "–",      # source unknown
        "–",      # company unknown
        "–",      # title unknown
        "–",      # location
        "–",      # region
        "–",      # date posted
        status,
        "–",      # url
        "–",      # resume
        "–",      # cover
        notes or "Created from Telegram button (job data pending next run)",
    ]
    ws.append(row)
    row_idx = ws.max_row

    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border    = CELL_BORDER
        cell.alignment = Alignment(wrap_text=False, vertical="center")
        cell.font      = Font(size=9)

    status_col = HEADERS.index("Status") + 1
    _color_status_cell(ws.cell(row=row_idx, column=status_col), status)

    wb.save(str(TRACKER_PATH))
    logger.info("Stub row created: job_id=%s  status=%s", job_id, status)
    return True


def get_applied_jobs() -> list:
    """Return all rows whose Status is 'Applied'."""
    return [j for j in get_all_jobs() if (j.get("Status") or "").lower() == "applied"]


def get_rejected_jobs() -> list:
    """Return all rows whose Status is 'Rejected'."""
    return [j for j in get_all_jobs() if (j.get("Status") or "").lower() == "rejected"]


def apply_status_overrides() -> int:
    """
    Read data/statuses.json (written by the dashboard UI) and apply any pending
    manual status changes to tracker.xlsx.
    Clears the file after applying.
    Returns the number of rows updated.
    """
    import json as _json
    statuses_path = TRACKER_PATH.parent / "statuses.json"
    if not statuses_path.exists():
        return 0
    try:
        overrides = _json.loads(statuses_path.read_text(encoding="utf-8"))
    except Exception:
        return 0
    if not overrides:
        return 0
    if not TRACKER_PATH.exists():
        return 0

    try:
        wb = openpyxl.load_workbook(str(TRACKER_PATH))
        ws = wb.active
        id_col     = HEADERS.index("ID") + 1
        status_col = HEADERS.index("Status") + 1
        applied = 0

        for row in ws.iter_rows(min_row=2):
            cell_id = row[id_col - 1].value
            if cell_id is not None:
                job_id = str(cell_id).strip()
                if job_id in overrides:
                    new_status = overrides[job_id]
                    row[status_col - 1].value = new_status
                    _color_status_cell(row[status_col - 1], new_status)
                    applied += 1

        if applied:
            wb.save(str(TRACKER_PATH))
            logger.info("Applied %d dashboard status override(s) from statuses.json", applied)
            # Clear applied overrides so they don't re-apply next run
            statuses_path.write_text("{}", encoding="utf-8")

        return applied
    except Exception as exc:
        logger.error("apply_status_overrides failed: %s", exc)
        return 0
